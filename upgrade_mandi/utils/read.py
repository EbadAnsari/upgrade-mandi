from concurrent.futures import Future, ProcessPoolExecutor, ThreadPoolExecutor
from threading import Thread
from typing import List

import pandas as pd

print("Modules Loaded")


start = 0
chunkSize = 100


def readChunk(
    file: str, sheetName: str, start: int, chunkSize: int, validColumnCount: int
) -> pd.DataFrame:
    print(validColumnCount)
    return pd.read_excel(
        file,
        skiprows=range(1, start),
        nrows=chunkSize,
        sheet_name=sheetName,
        usecols=[_ for _ in range(0, validColumnCount)],
    )


def readExcel(file: str, sheetName: str, maxRows: int, chunkSize: int) -> pd.DataFrame:
    _start = 0
    df = pd.read_excel(file, sheet_name=sheetName, nrows=1)
    validColumnCount = df.shape[1]
    df = pd.DataFrame()
    print(validColumnCount)
    noOfThreads = 4
    for _ in range(0, maxRows, chunkSize * noOfThreads):
        with ThreadPoolExecutor(max_workers=4) as worker:
            t: List[Future] = []
            print(f"saart: {_start}")
            for _ in range(noOfThreads):
                t.append(
                    worker.submit(
                        readChunk, file, sheetName, _start, chunkSize, validColumnCount
                    )
                )
                _start += chunkSize
            for f in t:
                result = f.result()
                if result.empty:
                    return df
                df = pd.concat([df, result], ignore_index=True)
    return df


if __name__ == "__main__":
    from openpyxl import load_workbook

    wb = load_workbook(
        "./../raw-sheets-dump/upgrade 25 aug 25 (2).xlsx", read_only=True
    )
    ws = wb.active
    print(
        readExcel(
            "./../raw-sheets-dump/upgrade 25 aug 25 (2).xlsx",
            "Sheet1",
            ws.max_row,
            100,
        ).columns
    )
    # print("Max possible rows:", ws.max_row)  # may include blanks
